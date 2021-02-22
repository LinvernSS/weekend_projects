import openpyxl
import logging
import datetime


def fix_data(workb):
    workb['VOC Rolling MoM']['B1'] = datetime.datetime(2018, 3, 1)
    workb['VOC Rolling MoM']['C1'] = datetime.datetime(2018, 2, 1)
    workb['VOC Rolling MoM']['D1'] = datetime.datetime(2018, 1, 1)
    return workb


def load_wb(fname):
    filepath = r'C:/Users/lucas/OneDrive/Desktop/Stuff/Job Stuff/Smoothstack git/weekend_proj/proj_1/xl_files/'
    try:
        workbook = openpyxl.load_workbook(filepath + fname)
    except IOError:
        logging.error('File {} not found.'.format(fname))
        return False
    else:
        logging.info('File {} loaded.'.format(fname))
        if type(workbook['VOC Rolling MoM']['B1']) != datetime.datetime:
            return fix_data(workbook)
        else:
            return workbook


def get_date(fname):
    fname = fname.split('_')
    month = fname[-2]
    month = month[0].upper() + month[1:]
    month = datetime.datetime.strptime(month, "%B").month
    year = int(fname[-1][:-5])
    return datetime.datetime(year, month, 1)


def get_data(date, book):
    trunc_date = lambda d: d.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    worksheets = []
    for sheet in book:
        worksheets.append(sheet)

    row_num = None
    for cell in worksheets[0]['A2':'A13']:  # retrieves row index of requested data
        val = trunc_date(cell[0].value)
        if date == val:
            row_num = str(cell[0].row)
            break

    if row_num is not None:
        get_worksheet0_data(row_num, worksheets[0])

        for cell in worksheets[1]['B1':'X1'][0]:  # retrieves column index of requested data
            val = trunc_date(cell.value)
            if date == val:
                col_name = cell.coordinate[0]
                break

        get_worksheet1_data(col_name, worksheets[1])

        return True
    else:
        logging.error('Requested date not found.')
        return False


def get_worksheet0_data(row_n, ws):
    for cell in ws['B' + row_n:'F' + row_n][0]:
        val = cell.value
        col = str(cell.coordinate[0])
        col_name = ws[col + '1'].value  # column label

        # can be either float or int
        if type(val) != int:
            val = str(round(val * 100, 2)) + '%'  # format floats into percentages
        else:
            val = f'{val:,}'  # adds commas in thousands place

        logging.info('{} : {}'.format(col_name.strip(), val))


def get_worksheet1_data(col_n, ws):
    # retrieves first section of data
    for cell in ws[col_n + '3':col_n + '9']:
        val = cell[0].value
        row = str(cell[0].row)
        temp_row_name = ws['A' + row].value  # row label

        # checks if first or second part of promoter scores
        if temp_row_name is not None:
            row_name = temp_row_name
            row_name = row_name.split()

            # checks if is actual promoter score or base size
            if len(row_name) > 2:
                row_name = row_name[0]
                if row_name == 'Promoters':
                    if val >= 200:
                        score = 'Good'
                    else:
                        score = 'Bad'
                elif row_name == 'Passives':
                    if val >= 100:
                        score = 'Good'
                    else:
                        score = 'Bad'
                else:
                    if val < 100:
                        score = 'Good'
                    else:
                        score = 'Bad'
                logging.info('Number of {} : {}, {}'.format(row_name, val, score))
            else:
                row_name = row_name[0] + ' ' + row_name[1]
                logging.info('{} : {}'.format(row_name, val))
        else:
            val = str(round(val * 100, 2)) + '%'
            logging.info('Percentage of {} : {}'.format(row_name, val))

    # retrieves second section of data
    for cell in ws[col_n + '12':col_n + '19']:
        val = cell[0].value
        row = str(cell[0].row)
        temp_row_name = ws['A' + row].value

        if temp_row_name is not None:
            temp_row_name = temp_row_name.split()

            # doesn't change row name if it is AARP total
            if len(temp_row_name) != 2:
                row_name = ' '.join(temp_row_name)
            else:
                val = str(round(val * 100, 2)) + '%'
                logging.info('{} : {}'.format(row_name, val))


if __name__ == "__main__":
    logging.basicConfig(filename='expediaData.log', level=logging.DEBUG)
    filenames = ['expedia_report_monthly_june_2017.xlsx',  # nonexistent file
                 'expedia_report_monthly_june_2018.xlsx',  # file 1 with nonexistent date
                 'expedia_report_monthly_january_2018.xlsx',  # file 1
                 'expedia_report_monthly_march_2018.xlsx',  # file 2
                 'expedia_report_monthly_march_2017.xlsx']  # file 2 with altered date

    for filename in filenames:
        wb = load_wb(filename)

        if type(wb) != bool:
            mon_yr = get_date(filename)
            if get_data(mon_yr, wb):
                logging.info('Data retrieval completed.\n')
            else:
                logging.info('Data retrieval failed.\n')
        else:
            logging.info('Data retrieval failed.\n')
